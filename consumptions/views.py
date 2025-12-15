from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
from .models import ConsumptionRegister, ConsumptionAccount, SubAccount
from .serializers import ConsumptionRegisterSerializer, ConsumptionAccountSerializer, SubAccountSerializer
from building_mgmt.models import Building
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import datetime
from decimal import Decimal


def get_building_from_request(request):
    """
    Extract building_id from request data or query params and return the Building object.
    Returns None if building_id is not provided or building doesn't exist.
    """
    building_id = request.data.get('building_id') or request.GET.get('building_id')
    if building_id:
        try:
            return Building.objects.get(id=building_id)
        except Building.DoesNotExist:
            return None
    return None


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def consumption_register(request):
    """
    GET: Retrieve consumption register entries with sub_account details.
         - Master role: sees all data for the selected building
         - Other roles: see only data they registered for the selected building
    POST: Create a new consumption register entry for the selected building.
    Expected POST data: {date, utilityType, value, building_id, subAccount (optional)}
    """
    if request.method == 'GET':
        building_id = request.GET.get('building_id')

        # Start with base queryset
        registers = ConsumptionRegister.objects.select_related('sub_account', 'building')

        # Filter by building if provided
        if building_id:
            registers = registers.filter(building_id=building_id)

        # Role-based filtering: master sees all, others see only their own
        if request.user.role != 'master':
            registers = registers.filter(created_by=request.user)

        serializer = ConsumptionRegisterSerializer(registers, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        # Validate building_id is provided
        building_id = request.data.get('building_id')
        if not building_id:
            return Response({
                'error': 'building_id is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Verify building exists
        try:
            building = Building.objects.get(id=building_id)
        except Building.DoesNotExist:
            return Response({
                'error': 'Building not found'
            }, status=status.HTTP_404_NOT_FOUND)

        serializer = ConsumptionRegisterSerializer(data=request.data)
        if serializer.is_valid():
            # Save with building and created_by
            serializer.save(building=building, created_by=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def consumption_register_detail(request, register_id):
    """
    GET: Retrieve a specific consumption register entry.
    PUT: Update a specific consumption register entry.
    DELETE: Delete a specific consumption register entry.
    Role-based access: master can access all, others only their own.
    """
    try:
        # Master role can access all registers, others only their own
        if request.user.role == 'master':
            register = ConsumptionRegister.objects.select_related('sub_account').get(id=register_id)
        else:
            register = ConsumptionRegister.objects.select_related('sub_account').get(
                id=register_id, created_by=request.user
            )
    except ConsumptionRegister.DoesNotExist:
        return Response({
            'error': 'Consumption register not found or you do not have permission to access it'
        }, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = ConsumptionRegisterSerializer(register)
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        serializer = ConsumptionRegisterSerializer(register, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response({
            'error': 'Invalid data',
            'details': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        register.delete()
        return Response({
            'message': 'Consumption register deleted successfully'
        }, status=status.HTTP_200_OK)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def consumption_account(request):
    """
    GET: Retrieve consumption account entries.
         - Master role: sees all data for the selected building
         - Other roles: see only data they registered for the selected building
    POST: Create a new consumption account entry for the selected building.
    Expected POST data: {amount, month, utilityType, building_id}
    """
    if request.method == 'GET':
        building_id = request.GET.get('building_id')

        # Start with base queryset
        accounts = ConsumptionAccount.objects.select_related('building')

        # Filter by building if provided
        if building_id:
            accounts = accounts.filter(building_id=building_id)

        # Role-based filtering: master sees all, others see only their own
        if request.user.role != 'master':
            accounts = accounts.filter(created_by=request.user)

        serializer = ConsumptionAccountSerializer(accounts, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        # Validate building_id is provided
        building_id = request.data.get('building_id')
        if not building_id:
            return Response({
                'error': 'building_id is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Verify building exists
        try:
            building = Building.objects.get(id=building_id)
        except Building.DoesNotExist:
            return Response({
                'error': 'Building not found'
            }, status=status.HTTP_404_NOT_FOUND)

        serializer = ConsumptionAccountSerializer(data=request.data)
        if serializer.is_valid():
            # Save with building and created_by
            serializer.save(building=building, created_by=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def consumption_account_detail(request, account_id):
    """
    GET: Retrieve a specific consumption account entry.
    PUT: Update a specific consumption account entry.
    DELETE: Delete a specific consumption account entry.
    Role-based access: master can access all, others only their own.
    """
    try:
        # Master role can access all accounts, others only their own
        if request.user.role == 'master':
            account = ConsumptionAccount.objects.get(id=account_id)
        else:
            account = ConsumptionAccount.objects.get(id=account_id, created_by=request.user)
    except ConsumptionAccount.DoesNotExist:
        return Response({
            'error': 'Consumption account not found or you do not have permission to access it'
        }, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = ConsumptionAccountSerializer(account)
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        serializer = ConsumptionAccountSerializer(account, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response({
            'error': 'Invalid data',
            'details': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        account.delete()
        return Response({
            'message': 'Consumption account deleted successfully'
        }, status=status.HTTP_200_OK)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def sub_account_list(request):
    """
    GET: Retrieve sub-accounts (can filter by utility_type and building_id query params).
         - Master role: sees all sub-accounts for the selected building
         - Other roles: see only sub-accounts they registered for the selected building
    POST: Create a new sub-account for the selected building.
    Expected POST data: {utilityType, name, building_id, icon (optional)}
    """
    if request.method == 'GET':
        building_id = request.GET.get('building_id')

        # Start with base queryset
        sub_accounts = SubAccount.objects.select_related('building')

        # Filter by building if provided
        if building_id:
            sub_accounts = sub_accounts.filter(building_id=building_id)

        # Filter by utility_type if provided
        utility_type = request.GET.get('utility_type')
        if utility_type:
            sub_accounts = sub_accounts.filter(utility_type=utility_type)

        # Role-based filtering: master sees all, others see only their own
        if request.user.role != 'master':
            sub_accounts = sub_accounts.filter(created_by=request.user)

        serializer = SubAccountSerializer(sub_accounts, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'POST':
        # Validate building_id is provided
        building_id = request.data.get('building_id')
        if not building_id:
            return Response({
                'error': 'building_id is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Verify building exists
        try:
            building = Building.objects.get(id=building_id)
        except Building.DoesNotExist:
            return Response({
                'error': 'Building not found'
            }, status=status.HTTP_404_NOT_FOUND)

        serializer = SubAccountSerializer(data=request.data)
        if serializer.is_valid():
            # Save with building and created_by
            serializer.save(building=building, created_by=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response({
            'error': 'Invalid data',
            'details': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def sub_account_detail(request, sub_account_id):
    """
    GET: Retrieve a specific sub-account.
    PUT: Update a specific sub-account.
    DELETE: Delete a specific sub-account.
    Role-based access: master can access all, others only their own.
    """
    try:
        # Master role can access all sub-accounts, others only their own
        if request.user.role == 'master':
            sub_account = SubAccount.objects.get(id=sub_account_id)
        else:
            sub_account = SubAccount.objects.get(id=sub_account_id, created_by=request.user)
    except SubAccount.DoesNotExist:
        return Response({
            'error': 'Sub-account not found or you do not have permission to access it'
        }, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = SubAccountSerializer(sub_account)
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        serializer = SubAccountSerializer(sub_account, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response({
            'error': 'Invalid data',
            'details': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        sub_account.delete()
        return Response({
            'message': 'Sub-account deleted successfully'
        }, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_consumption_excel(request):
    """
    Export consumption register data to Excel file.
    Optional query parameters: utility_type (water, electricity, gas), building_id
    Role-based access: master sees all, others only their own.
    """
    # Get query parameters
    utility_type = request.GET.get('utility_type', None)
    building_id = request.GET.get('building_id', None)

    # Get consumption registers with related data
    registers = ConsumptionRegister.objects.select_related('sub_account', 'building')

    # Filter by building if provided
    if building_id:
        registers = registers.filter(building_id=building_id)

    # Role-based filtering: master sees all, others see only their own
    if request.user.role != 'master':
        registers = registers.filter(created_by=request.user)

    # Filter by utility type if provided
    if utility_type:
        registers = registers.filter(utility_type=utility_type)

    # Order by date ascending (oldest first)
    registers = registers.order_by('date')

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active

    # Set title based on filter
    if utility_type:
        utility_display = dict(ConsumptionRegister.UTILITY_TYPE_CHOICES).get(utility_type, utility_type)
        ws.title = f"{utility_display} Consumption Report"
        report_title = f"{utility_display} Consumption Report"
    else:
        ws.title = "Consumption Report"
        report_title = "Consumption Report - All Utilities"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = report_title
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")

    # Export date
    ws.merge_cells('A2:F2')
    date_info = ws['A2']
    date_info.value = f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    date_info.alignment = Alignment(horizontal="center")

    # Headers
    headers = ['Date', 'Utility Type', 'Sub Account', 'Value', 'Consumption', 'Unit']

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Write consumption data with daily change calculation
    row = 5
    total_value = Decimal('0')
    total_consumption = Decimal('0')
    previous_value = None

    # Convert queryset to list for easier indexing
    registers_list = list(registers)

    for index, register in enumerate(registers_list):
        # Get sub-account name
        sub_account_name = register.sub_account.name if register.sub_account else "N/A"

        # Map utility_type to readable text
        utility_display = dict(ConsumptionRegister.UTILITY_TYPE_CHOICES).get(
            register.utility_type,
            register.utility_type
        )

        # Get unit label based on utility type
        unit_map = {
            'water': 'm³',
            'electricity': 'kWh',
            'gas': 'm³'
        }
        unit_label = unit_map.get(register.utility_type, '')

        # Parse value and add to total
        value_decimal = Decimal(str(register.value))
        total_value += value_decimal

        # Calculate consumption (difference from previous day)
        consumption = None
        if index > 0:
            previous_value = Decimal(str(registers_list[index - 1].value))
            consumption = value_decimal - previous_value
            total_consumption += consumption

        # Format value to remove trailing zeros
        value_formatted = float(value_decimal)
        consumption_formatted = float(consumption) if consumption is not None else None

        data = [
            register.date.strftime('%Y-%m-%d'),
            utility_display,
            sub_account_name,
            value_formatted,
            consumption_formatted if consumption_formatted is not None else '',
            unit_label
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = border
            # Center alignment for all cells
            cell.alignment = Alignment(horizontal="center")

        row += 1

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0

        # Check header length
        max_length = max(max_length, len(str(headers[col-1])))

        # Check data lengths
        for row_num in range(5, row):
            cell_value = ws.cell(row=row_num, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        # Set column width with some padding
        adjusted_width = min(max_length + 2, 50)  # Max width of 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # Add summary information
    summary_row = row + 2
    ws.merge_cells(f'A{summary_row}:C{summary_row}')
    summary_cell = ws[f'A{summary_row}']
    summary_cell.value = "Total:"
    summary_cell.font = Font(bold=True)
    summary_cell.alignment = Alignment(horizontal="right")

    # Total value
    total_cell = ws.cell(row=summary_row, column=4)
    total_cell.value = float(total_value)
    total_cell.font = Font(bold=True)
    total_cell.alignment = Alignment(horizontal="center")
    total_cell.border = border

    # Total consumption
    total_consumption_cell = ws.cell(row=summary_row, column=5)
    total_consumption_cell.value = float(total_consumption)
    total_consumption_cell.font = Font(bold=True)
    total_consumption_cell.alignment = Alignment(horizontal="center")
    total_consumption_cell.border = border

    # Unit for total
    unit_cell = ws.cell(row=summary_row, column=6)
    if utility_type:
        unit_cell.value = unit_map.get(utility_type, '')
    else:
        unit_cell.value = "Mixed"
    unit_cell.font = Font(bold=True)
    unit_cell.alignment = Alignment(horizontal="center")
    unit_cell.border = border

    # Record count
    summary_row += 1
    ws.merge_cells(f'A{summary_row}:F{summary_row}')
    count_cell = ws[f'A{summary_row}']
    count_cell.value = f"Total Records: {len(registers_list)}"
    count_cell.font = Font(bold=True)
    count_cell.alignment = Alignment(horizontal="center")

    # Create HTTP response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # Generate filename
    if utility_type:
        filename = f"consumption_{utility_type}_report.xlsx"
    else:
        filename = "consumption_report.xlsx"

    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_consumption_excel(request):
    """
    Import consumption register data from Excel file.
    Required: building_id in request data.
    Expected file format matches the export format:
    - Row 1: Title
    - Row 2: Export date
    - Row 4: Headers (Date, Utility Type, Sub Account, Value, Unit)
    - Row 5+: Data rows
    """
    # Validate building_id is provided
    building_id = request.data.get('building_id') or request.GET.get('building_id')
    if not building_id:
        return Response({
            'error': 'building_id is required'
        }, status=status.HTTP_400_BAD_REQUEST)

    # Verify building exists
    try:
        building = Building.objects.get(id=building_id)
    except Building.DoesNotExist:
        return Response({
            'error': 'Building not found'
        }, status=status.HTTP_404_NOT_FOUND)

    # Check if file was uploaded
    if 'file' not in request.FILES:
        return Response({
            'error': 'No file uploaded. Please upload an Excel file.'
        }, status=status.HTTP_400_BAD_REQUEST)

    excel_file = request.FILES['file']

    # Validate file type
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return Response({
            'error': 'Invalid file type. Please upload an Excel file (.xlsx or .xls).'
        }, status=status.HTTP_400_BAD_REQUEST)

    try:
        # Load workbook
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
    except Exception as e:
        return Response({
            'error': f'Error reading Excel file: {str(e)}'
        }, status=status.HTTP_400_BAD_REQUEST)

    try:
        # Find header row (should be row 4 based on export format)
        header_row = 4
        headers = []
        for col in range(1, 7):  # 6 columns expected now (added Consumption)
            try:
                cell_value = ws.cell(row=header_row, column=col).value
                if cell_value:
                    headers.append(str(cell_value).strip())
                else:
                    headers.append("")
            except Exception:
                headers.append("")

        # Expected headers from export
        expected_headers = ['Date', 'Utility Type', 'Sub Account', 'Value', 'Consumption', 'Unit']

        # Validate headers - Date, Utility Type, Sub Account, and Value are required
        if len(headers) < 4:
            return Response({
                'error': 'Invalid Excel format. Missing required columns.',
                'expected_headers': expected_headers,
                'found_headers': headers
            }, status=status.HTTP_400_BAD_REQUEST)

        # Parse data rows (starting from row 5)
        consumption_data = []
        errors = []
        warnings = []
        row_num = 5

        # Create reverse mappings for choices
        utility_type_map = {v: k for k, v in ConsumptionRegister.UTILITY_TYPE_CHOICES}

        # Get sub-accounts for the building to map names to IDs
        sub_accounts_by_name = {}
        for sub_account in SubAccount.objects.filter(building=building):
            key = f"{sub_account.utility_type}_{sub_account.name}"
            sub_accounts_by_name[key] = sub_account

        # Pre-fetch existing consumption registers for this building by date and utility type to avoid duplicates
        existing_registers = {}
        for register in ConsumptionRegister.objects.filter(building=building):
            key = f"{register.date}_{register.utility_type}_{register.sub_account_id if register.sub_account else 'none'}"
            existing_registers[key] = register

        while True:
            try:
                # Read row data (Consumption column is skipped - it's auto-calculated)
                date_str = ws.cell(row=row_num, column=1).value
                utility_type_str = ws.cell(row=row_num, column=2).value
                sub_account_str = ws.cell(row=row_num, column=3).value
                value_str = ws.cell(row=row_num, column=4).value
                # Column 5 (Consumption) is ignored during import as it's calculated automatically

                # Stop if we hit an empty row or summary section
                if not date_str or str(date_str).strip() == '' or 'Total' in str(date_str):
                    break

                # Parse date
                if isinstance(date_str, datetime):
                    date_obj = date_str.date()
                else:
                    try:
                        date_obj = datetime.strptime(str(date_str).strip(), '%Y-%m-%d').date()
                    except ValueError:
                        errors.append(f"Row {row_num}: Invalid date format '{date_str}'. Expected YYYY-MM-DD.")
                        row_num += 1
                        continue

                # Parse utility type
                utility_type_clean = str(utility_type_str).strip() if utility_type_str else ''
                utility_type = utility_type_map.get(utility_type_clean)

                if not utility_type:
                    # Try lowercase match
                    utility_type = utility_type_clean.lower()
                    if utility_type not in ['water', 'electricity', 'gas']:
                        errors.append(f"Row {row_num}: Invalid utility type '{utility_type_str}'")
                        row_num += 1
                        continue

                # Parse sub-account
                sub_account_obj = None
                if sub_account_str and str(sub_account_str).strip() != 'N/A':
                    sub_account_name = str(sub_account_str).strip()
                    key = f"{utility_type}_{sub_account_name}"
                    sub_account_obj = sub_accounts_by_name.get(key)

                    if not sub_account_obj:
                        warnings.append(f"Row {row_num}: Sub-account '{sub_account_name}' not found for {utility_type}. Record will be created without sub-account.")

                # Parse value
                try:
                    if isinstance(value_str, (int, float)):
                        value = str(value_str)
                    else:
                        value = str(value_str).strip()

                    # Validate it's a number
                    float(value)
                except (ValueError, TypeError):
                    errors.append(f"Row {row_num}: Invalid value '{value_str}'. Must be a number.")
                    row_num += 1
                    continue

                # Check if this record already exists
                sub_account_id = sub_account_obj.id if sub_account_obj else 'none'
                check_key = f"{date_obj}_{utility_type}_{sub_account_id}"

                if check_key in existing_registers:
                    # Update existing record
                    existing_register = existing_registers[check_key]
                    existing_register.value = value
                    consumption_data.append({
                        'register': existing_register,
                        'is_update': True
                    })
                else:
                    # Create new record with building and created_by
                    new_register = ConsumptionRegister(
                        date=date_obj,
                        utility_type=utility_type,
                        sub_account=sub_account_obj,
                        value=value,
                        building=building,
                        created_by=request.user
                    )
                    consumption_data.append({
                        'register': new_register,
                        'is_update': False
                    })

                row_num += 1

            except Exception as e:
                errors.append(f"Row {row_num}: Unexpected error - {str(e)}")
                row_num += 1
                # Stop if we've read too many rows (safety check)
                if row_num > 10000:
                    break

        # If there are critical errors, return them
        if errors and len(consumption_data) == 0:
            return Response({
                'error': 'Failed to import any records',
                'errors': errors,
                'warnings': warnings
            }, status=status.HTTP_400_BAD_REQUEST)

        # Save consumption data
        created_count = 0
        updated_count = 0
        save_errors = []

        # Separate new records and updates
        new_registers = [item['register'] for item in consumption_data if not item['is_update']]
        update_registers = [item['register'] for item in consumption_data if item['is_update']]

        # Bulk create new registers
        if new_registers:
            try:
                ConsumptionRegister.objects.bulk_create(new_registers, batch_size=100)
                created_count = len(new_registers)
            except Exception as e:
                # Fallback to individual saves
                for register in new_registers:
                    try:
                        register.save()
                        created_count += 1
                    except Exception as save_error:
                        save_errors.append(f"Error saving record for {register.date}: {str(save_error)}")

        # Update existing registers
        for register in update_registers:
            try:
                register.save()
                updated_count += 1
            except Exception as save_error:
                save_errors.append(f"Error updating record for {register.date}: {str(save_error)}")

        # Prepare response
        response_data = {
            'message': 'Import completed',
            'summary': {
                'created': created_count,
                'updated': updated_count,
                'total': created_count + updated_count,
                'errors': len(save_errors)
            }
        }

        if warnings:
            response_data['warnings'] = warnings

        if save_errors:
            response_data['save_errors'] = save_errors

        if errors:
            response_data['parsing_errors'] = errors

        return Response(response_data, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({
            'error': f'Unexpected error during import: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)